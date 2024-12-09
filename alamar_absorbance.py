import openpyxl
import numpy as np
from tabulate import tabulate
from docx import Document
from openpyxl.utils.cell import coordinate_from_string
import xlwings as xw
import tkinter as tk
from tkinter import filedialog
import os

def get_replicate_cells_from_range(cell_range):
    """
    Processes a range of cells to extract replicate pairs for 570 nm and 600 nm readings.

    Parameters:
    - cell_range (str): The Excel cell range selected by the user (e.g., 'A1:A8').

    Returns:
    - replicate_cells (list of tuples): A list where each tuple contains the cell addresses
      for the 570 nm and 600 nm readings of a replicate.
    """
    # Split the cell range into start and end cells
    start_cell, end_cell = cell_range.split(':')

    # Extract column letters and row numbers from the cell addresses
    start_col, start_row = coordinate_from_string(start_cell)
    end_col, end_row = coordinate_from_string(end_cell)
    start_row = int(start_row)
    end_row = int(end_row)

    replicate_cells = []
    # Iterate over the rows in steps of 2 since each replicate consists of two readings
    for row in range(start_row, end_row + 1, 2):
        cell_570 = f"{start_col}{row}"         # Cell for 570 nm reading
        cell_600 = f"{start_col}{row + 1}"     # Cell for 600 nm reading
        replicate_cells.append((cell_570, cell_600))
    return replicate_cells

def process_replicates(sheet, replicate_cells):
    """
    Processes the replicate cells to calculate adjusted absorbance values.

    Parameters:
    - sheet (Worksheet): The worksheet object from openpyxl containing the data.
    - replicate_cells (list of tuples): List of cell address pairs for each replicate.

    Returns:
    - replicate_values (list): List of calculated adjusted absorbance values for each replicate.
    """
    replicate_values = []
    for cell_570, cell_600 in replicate_cells:
        # Retrieve absorbance values from the specified cells
        value_570 = sheet[cell_570].value
        value_600 = sheet[cell_600].value

        if value_570 is not None and value_600 is not None:
            # Calculate the adjusted absorbance using the specified formula
            adjusted_absorbance = (value_570 * 117216) - (value_600 * 80586)
            replicate_values.append(adjusted_absorbance)
        else:
            print(f"Warning: Missing data in cells {cell_570} or {cell_600}. Skipping these cells.")
    return replicate_values

def main():
    """
    Main function to execute the alamarBlue assay data processing script.
    It guides the user through selecting the data file, inputting sample information,
    selecting data ranges, and calculates cell viability percentages.
    """
    # Use tkinter to open a file dialog and select the Excel file
    print("Select the Excel file...")
    root = tk.Tk()
    root.withdraw()  # Hide the root window
    file_path = filedialog.askopenfilename(
        title="Select the Excel file",
        filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")]
    )
    if not file_path:
        print("No file selected. Exiting.")
        return

    # Load the workbook using openpyxl
    try:
        workbook = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return

    # Prompt the user for the sheet name
    print(f"Available sheets: {workbook.sheetnames}")
    sheet_name = input("Enter the sheet name: ")
    if sheet_name not in workbook.sheetnames:
        print(f"Error: Sheet '{sheet_name}' does not exist in the workbook.")
        return
    sheet = workbook[sheet_name]

    # Process the positive control sample
    positive_control_name = input("Enter the positive control sample name: ")

    print(f"Please select the range of cells for {positive_control_name} in Excel and then press Enter here.")
    # Open Excel and let the user select the data range for the positive control
    app = xw.App(visible=True)
    wb = app.books.open(file_path)
    sheet_xl = wb.sheets[sheet_name]
    sheet_xl.activate()
    input("After selecting the range in Excel, return here and press Enter to continue...")

    # Get the selected range from Excel
    try:
        selected_range = app.selection.address.replace('$', '')  # Remove dollar signs from the address
        replicate_cells = get_replicate_cells_from_range(selected_range)
        positive_control_replicate_values = process_replicates(sheet, replicate_cells)
        wb.close()
        app.quit()
    except Exception as e:
        print(f"Error getting selected range: {e}")
        wb.close()
        app.quit()
        return

    if not positive_control_replicate_values:
        print("No valid data found for positive control. Exiting.")
        return

    # Calculate the average adjusted absorbance for the positive control
    positive_control_average = np.mean(positive_control_replicate_values)
    positive_control_viability_percentage = 100.0  # Set positive control viability to 100%

    # Prompt the user about outlier removal for the samples
    remove_outliers = input("Do you want to remove outliers for the samples? (yes/no): ").strip().lower() == "yes"

    # Prompt for the number of additional samples to process
    num_samples = int(input("Enter the number of additional samples: "))
    samples_data = []

    # Process each additional sample
    for _ in range(num_samples):
        sample_name = input("Enter the sample name: ")

        print(f"Please select the range of cells for {sample_name} in Excel and then press Enter here.")
        # Open Excel and let the user select the data range for the sample
        app = xw.App(visible=True)
        wb = app.books.open(file_path)
        sheet_xl = wb.sheets[sheet_name]
        sheet_xl.activate()
        input("After selecting the range in Excel, return here and press Enter to continue...")

        # Get the selected range from Excel
        try:
            selected_range = app.selection.address.replace('$', '')
            replicate_cells = get_replicate_cells_from_range(selected_range)
            replicate_values = process_replicates(sheet, replicate_cells)
            wb.close()
            app.quit()
        except Exception as e:
            print(f"Error getting selected range: {e}")
            wb.close()
            app.quit()
            continue  # Skip to the next sample if an error occurs

        if not replicate_values:
            print(f"No valid data found for {sample_name}. Skipping.")
            continue

        samples_data.append({
            'sample_name': sample_name,
            'replicate_values': replicate_values
        })

    # Calculate viability percentages and prepare the results table
    table_data = [["Sample", "Cell Viability %"]]
    table_data.append([positive_control_name, positive_control_viability_percentage])

    for sample in samples_data:
        sample_name = sample['sample_name']
        replicate_values = np.array(sample['replicate_values'])

        # Remove outliers if applicable
        if remove_outliers and len(replicate_values) > 2:
            # Calculate interquartile range (IQR)
            Q1 = np.percentile(replicate_values, 25)
            Q3 = np.percentile(replicate_values, 75)
            IQR = Q3 - Q1
            # Determine outlier thresholds
            Tmin = Q1 - (1.5 * IQR)
            Tmax = Q3 + (1.5 * IQR)
            # Filter out the outliers
            filtered_values = replicate_values[(replicate_values >= Tmin) & (replicate_values <= Tmax)]
            if len(filtered_values) != len(replicate_values):
                print(f"Removed outliers for {sample_name}: {set(replicate_values) - set(filtered_values)}")
        else:
            filtered_values = replicate_values

        # Calculate the average adjusted absorbance for the sample
        average = np.mean(filtered_values)
        # Calculate the cell viability percentage relative to the positive control
        viability_percentage = round((average / positive_control_average) * 100, 1)
        table_data.append([sample_name, viability_percentage])

    # Display the results table
    print("\nCell Viability Results:")
    print(tabulate(table_data, headers="firstrow", tablefmt="grid"))

    # Optionally export the results table to a Word document
    if input("Do you want to export the table to a Word document? (yes/no): ").strip().lower() == "yes":
        try:
            document = Document()
            table = document.add_table(rows=len(table_data), cols=len(table_data[0]))
            for i, row in enumerate(table_data):
                hdr_cells = table.rows[i].cells
                for j, cell_value in enumerate(row):
                    hdr_cells[j].text = str(cell_value)
            output_file = input("Enter the output file name (with .docx extension): ")
            if not output_file.endswith('.docx'):
                output_file += '.docx'
            document.save(output_file)
            print(f"Word document saved successfully as '{output_file}'.")
        except Exception as e:
            print(f"An error occurred while exporting to Word document: {e}")

if __name__ == "__main__":
    main()
